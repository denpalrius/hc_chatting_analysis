from io import BytesIO
from datetime import datetime
import streamlit as st
import pandas as pd
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def get_individual_name(content: bytes, filename: str) -> str:
    """Extract the person's name from D3 (xls or xlsx)."""
    if filename.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=content)
        sh = wb.sheet_by_index(0)
        return sh.cell_value(2, 3).split(",")[0].strip()
    else:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        raw = wb.worksheets[0].cell(row=3, column=4).value
        return str(raw).split(",")[0].strip()


def find_all_sections(raw: pd.DataFrame) -> list[int]:
    """
    Return a sorted list of all row-indices where column A == "Date" (case‐insensitive),
    but only those that appear below the "Time Zone:" row.
    """
    col0 = raw.iloc[:, 0].astype(str)
    # find index of "Time Zone:" (so that we skip any 'Date' cells above it)
    tz_mask = col0.str.contains("Time Zone:", case=False, na=False)
    tz_idx = tz_mask.idxmax() if tz_mask.any() else -1

    # find all rows where column A literally equals "Date"
    date_mask = col0.str.strip().str.lower() == "date"
    all_date_rows = [r for r in raw.index[date_mask] if r > tz_idx]
    return sorted(all_date_rows)


def find_section_end(raw: pd.DataFrame, header_row: int) -> int:
    """
    Given a header_row (where column A == "Date"), find the first row > header_row
    in which column C (zero-based index 2) literally equals "Total" (any case).
    If none, return len(raw).
    """
    col2 = raw.iloc[:, 2].astype(str).str.strip().str.lower()
    # look for “total” in column C below header_row
    total_mask = (col2 == "total") & (raw.index > header_row)
    candidates = raw.index[total_mask]
    return int(candidates[0]) if len(candidates) > 0 else len(raw)


def get_acronym(name: str) -> str:
    """Return the acronym (initials) for a given name string."""
    return "".join([part[0].upper() for part in name.split() if part])


def parse_file(content: bytes, filename: str) -> pd.DataFrame:
    """
    Read one per‐person sheet, find both the Oversight and the PDN‐LPN blocks
    (or any other ISP section), and return a DataFrame with columns:
      [Date, Service Provider, Individual, Duration_hours (decimal)].
    """
    indiv_full = get_individual_name(content, filename)
    indiv = get_acronym(indiv_full)

    # read entire sheet without headers
    raw = pd.read_excel(BytesIO(content), header=None)

    # find every header row where column A == "Date" (below "Time Zone:")
    header_rows = find_all_sections(raw)
    parsed_dfs: list[pd.DataFrame] = []

    for hdr in header_rows:
        end_row = find_section_end(raw, hdr)
        section = raw.iloc[hdr + 1 : end_row]

        # try to parse column A as a date; invalid dates become NaT and get dropped
        parsed_dates = pd.to_datetime(section.iloc[:, 0], errors="coerce")
        valid_mask = parsed_dates.notna()
        if not valid_mask.any():
            continue

        # We know column E (zero-based index 4) is Duration (minutes),
        # and column G (zero-based index 6) is Service Provider.
        df_block = pd.DataFrame(
            {
                "Date": parsed_dates[valid_mask].dt.date,
                "Duration_min": section.iloc[:, 4][valid_mask].astype(float),
                "Service Provider": section.iloc[:, 6][valid_mask].astype(str),
            }
        )
        # convert minutes → decimal hours
        df_block["Duration_hours"] = df_block["Duration_min"] / 60.0
        df_block["Individual"] = indiv

        # keep only the columns we need
        parsed_dfs.append(
            df_block[["Date", "Service Provider", "Individual", "Duration_hours"]]
        )

    if len(parsed_dfs) == 0:
        # If for some reason no section was parsed, return an empty DataFrame
        return pd.DataFrame(
            columns=["Date", "Service Provider", "Individual", "Duration_hours"]
        )

    return pd.concat(parsed_dfs, ignore_index=True)


def build_summary_workbook(df_raw: pd.DataFrame) -> Workbook:
    """
    Build a 'DailyMatrix' sheet in a new Workbook.  For each date:
      - list each Service Provider in column A,
      - list each Individual in columns B..,
      - write the total hours (decimal) for that provider‐individual as a numeric cell,
      - then add a SUM-formula for the provider’s row,
      - then a “Total hours for individual” row that sums down each column,
      - then a “24hr cap remaining” row as =24 - [that total].
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "DailyMatrix"
    bold = Font(bold=True)

    # sort list of Individuals
    individuals = sorted(df_raw["Individual"].unique())
    total_col = (
        len(individuals) + 2
    )  # column 1 = Provider, columns 2..N+1 = Individuals, column N+2 = Provider Total

    row = 1
    for date in sorted(df_raw["Date"].unique()):
        day_df = df_raw[df_raw["Date"] == date]

        # Date header
        ws.cell(row=row, column=1, value=date.strftime("%m/%d/%Y")).font = bold
        row += 1

        # Column headers: "Service Provider" | Individual names … | "Provider Total"
        ws.cell(row=row, column=1, value="Service Provider").font = bold
        for idx, name in enumerate(individuals, start=2):
            ws.cell(row=row, column=idx, value=name).font = bold
        ws.cell(row=row, column=total_col, value="Provider Total").font = bold
        row += 1

        provider_start = row

        # For each Service Provider on that date, write one row
        for prov in day_df["Service Provider"].unique():
            ws.cell(row=row, column=1, value=prov)

            # For each individual, sum up all decimal hours for (prov,individual)
            for idx, name in enumerate(individuals, start=2):
                hrs = day_df[
                    (day_df["Service Provider"] == prov)
                    & (day_df["Individual"] == name)
                ]["Duration_hours"].sum()
                # write integer hours if whole, otherwise 2-dp decimal
                if hrs.is_integer():
                    cell = ws.cell(row=row, column=idx, value=int(hrs))
                else:
                    val = round(hrs, 2)
                    cell = ws.cell(row=row, column=idx, value=val)
                    cell.number_format = "0.00"

            # In the “Provider Total” column, write a SUM formula across that row
            first_data_col = get_column_letter(2)
            last_data_col = get_column_letter(1 + len(individuals))
            ws.cell(
                row=row,
                column=total_col,
                value=f"=SUM({first_data_col}{row}:{last_data_col}{row})",
            ).font = bold

            row += 1

        provider_end = row - 1

        # Row: “Total hours for individual”  – sum down each individual column
        ws.cell(row=row, column=1, value="Total hours for individual").font = bold
        for idx, _ in enumerate(individuals, start=2):
            col_letter = get_column_letter(idx)
            ws.cell(
                row=row,
                column=idx,
                value=f"=SUM({col_letter}{provider_start}:{col_letter}{provider_end})",
            ).font = bold
        row += 1

        # Row: “Total hrs pending in a 24hr period” = 24 – [that total] for each individual
        ws.cell(row=row, column=1, value="Total hrs pending in a 24hr period").font = (
            bold
        )
        for idx in range(2, 2 + len(individuals)):
            col_letter = get_column_letter(idx)
            # subtract the total‐hours cell (just above) from 24
            ws.cell(row=row, column=idx, value=f"=24 - {col_letter}{row-1}").font = bold
        row += 2  # leave one blank row before next date block

    return wb


def main():
    st.title("Daily Staffing Analysis App")

    files = st.file_uploader(
        "Upload your per-person Excel files (.xls/.xlsx).",
        type=["xls", "xlsx"],
        accept_multiple_files=True,
    )
    if not files:
        return

    # parse each file into a DataFrame, then append them
    dfs = [parse_file(f.read(), f.name) for f in files]
    df_raw = pd.concat(dfs, ignore_index=True)

    if df_raw.empty:
        st.error("No valid 'Date' blocks were found in the uploaded files.")
        return

    wb = build_summary_workbook(df_raw)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    st.success("✅ Staffing summary is ready!")
    today = datetime.now().strftime("%Y-%m-%d")
    st.download_button(
        "Download Summary Excel",
        data=buf,
        file_name=f"daily_summary_{today}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
