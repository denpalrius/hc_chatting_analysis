import io
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelFormatter:
    """
    Handles Excel file formatting and output generation
    """
    
    def __init__(self):
        self.colors = {
            'RED_FILL': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),
            'YELLOW_FILL': PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid'),
            'GREEN_FILL': PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid'),
            'ORANGE_FILL': PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid'),
            'GREEN_FONT': Font(color='FF00B050'),
            'BOLD_FONT': Font(bold=True)
        }
    
    def create_formatted_excel(self, workbook, changes_log):
        """
        Create a properly formatted Excel file from the processed workbook
        """
        try:
            # Create a summary sheet with changes log
            self._add_summary_sheet(workbook, changes_log)
            
            # Save to bytes for download
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            st.error(f"Error creating formatted Excel file: {str(e)}")
            return None
    
    def _add_summary_sheet(self, workbook, changes_log):
        """
        Add a summary sheet showing all changes made
        """
        try:
            # Create summary sheet
            if 'Summary' in workbook.sheetnames:
                summary_ws = workbook['Summary']
            else:
                summary_ws = workbook.create_sheet('Summary', 0)  # Insert at beginning
            
            # Clear existing content
            summary_ws.delete_rows(1, summary_ws.max_row)
            
            # Add header
            headers = ['Change Type', 'Date', 'Provider', 'Individual', 'Details']
            for col_num, header in enumerate(headers, 1):
                cell = summary_ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = self.colors['BOLD_FONT']
                cell.fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid')
            
            # Add changes log
            row_num = 2
            for log_entry in changes_log:
                # Parse log entry (this is a simple implementation)
                parts = log_entry.split(': ', 1)
                if len(parts) == 2:
                    date_part = parts[0]
                    details = parts[1]
                    
                    # Try to extract more details
                    change_type = 'Modified'
                    provider = 'Unknown'
                    individual = 'Unknown'
                    
                    if 'Added' in details:
                        change_type = 'Added'
                    elif 'Increased' in details:
                        change_type = 'Increased'
                    elif 'Reduced' in details:
                        change_type = 'Reduced'
                    elif 'Set' in details:
                        change_type = 'Set'
                    
                    # Basic parsing for provider and individual
                    words = details.split()
                    for i, word in enumerate(words):
                        if word in ['DD', 'DM', 'OT']:
                            individual = word
                            break
                    
                    # Add row
                    summary_ws.cell(row=row_num, column=1, value=change_type)
                    summary_ws.cell(row=row_num, column=2, value=date_part)
                    summary_ws.cell(row=row_num, column=3, value=provider)
                    summary_ws.cell(row=row_num, column=4, value=individual)
                    summary_ws.cell(row=row_num, column=5, value=details)
                    
                    row_num += 1
            
            # Auto-adjust column widths
            for column in summary_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                summary_ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            st.warning(f"Could not create summary sheet: {str(e)}")
    
    def preserve_original_formatting(self, original_wb, processed_wb):
        """
        Preserve original formatting while maintaining new changes
        This is a complex operation that would need careful implementation
        """
        # This is a placeholder for more sophisticated formatting preservation
        # In a full implementation, we would:
        # 1. Copy original formatting
        # 2. Apply new formatting only where changes were made
        # 3. Preserve formulas and cell references
        pass
    
    def apply_color_coding_rules(self, worksheet, changes_made):
        """
        Apply color coding according to the business rules
        """
        try:
            # This would implement the specific color coding rules
            # based on the types of changes made
            for change in changes_made:
                row = change.get('row')
                col = change.get('col')
                change_type = change.get('type')
                
                if row and col and change_type:
                    cell = worksheet.cell(row=row, column=col)
                    
                    if change_type == 'new_addition':
                        cell.fill = self.colors['GREEN_FILL']
                    elif change_type == 'modified_positive':
                        cell.fill = self.colors['GREEN_FILL']
                    elif change_type == 'reduced':
                        cell.fill = self.colors['ORANGE_FILL']
                    elif change_type == 'existing_modified':
                        cell.fill = self.colors['YELLOW_FILL']
                        
        except Exception as e:
            st.warning(f"Error applying color coding: {str(e)}")
    
    def validate_excel_output(self, workbook):
        """
        Validate that the output Excel file meets requirements
        """
        try:
            issues = []
            
            # Check if main sheet exists
            if 'DailyMatrix' not in workbook.sheetnames:
                issues.append("Main 'DailyMatrix' sheet not found")
            
            # Check if summary sheet was created
            if 'Summary' not in workbook.sheetnames:
                issues.append("Summary sheet was not created")
            
            # Additional validation can be added here
            
            return issues
            
        except Exception as e:
            return [f"Error validating output: {str(e)}"]
