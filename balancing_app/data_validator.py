import pandas as pd
import streamlit as st
import openpyxl
from datetime import datetime

class DataValidator:
    """
    Validates the structure and content of healthcare schedule Excel files
    """
    
    def __init__(self):
        self.expected_individuals = []  # Will be extracted from file
        self.supplemental_providers = [
            'Charles Sagini, RN/House Manager',
            'Josephine Sagini, RN/Program Manager', 
            'Faith Murerwa, RN/House Supervisor'
        ]
        self.additional_providers = []  # Will be populated by user input
        
        # Formatting constants based on analysis
        self.colors = {
            'RED_FILL': 'FFFF0000',      # Date cells
            'YELLOW_FILL': 'FFFFC000',   # Modified entries  
            'GREEN_FONT': 'FF00B050',    # New providers
            'ORANGE_FILL': 'FFFFA500',   # Reduced entries (to be added)
            'GREEN_FILL': 'FF00FF00'     # New additions (to be added)
        }
    
    def validate_file_structure(self, excel_sheets):
        """
        Validate that the Excel file has the expected structure
        """
        try:
            if not excel_sheets:
                st.error("No sheets found in Excel file")
                return False
            
            # Look for DailyMatrix sheet or use the first sheet
            if 'DailyMatrix' in excel_sheets:
                main_sheet = 'DailyMatrix'
            else:
                main_sheet = list(excel_sheets.keys())[0]
                st.warning(f"Expected 'DailyMatrix' sheet not found. Using '{main_sheet}' instead.")
            
            df = excel_sheets[main_sheet]
            
            if df.empty:
                st.error("The main sheet is empty")
                return False
            
            if df.shape[1] < 5:
                st.error("Expected at least 5 columns (Service Provider, DD, DM, OT, Provider Total)")
                return False
            
            return True
            
        except Exception as e:
            st.error(f"Error validating file structure: {str(e)}")
            return False
    
    def is_schedule_sheet(self, df):
        """
        Check if a dataframe represents a schedule sheet
        """
        try:
            df_str = df.astype(str)
            
            # Check for presence of individuals (DD, DM, OT)
            has_individuals = any(
                any(individual in str(cell) for cell in row) 
                for _, row in df_str.iterrows() 
                for individual in self.expected_individuals
            )
            
            # Check for dates pattern (MM/DD/YYYY)
            has_dates = any(
                self._is_date_string(str(cell))
                for _, row in df_str.iterrows()
                for cell in row
            )
            
            return has_individuals and has_dates
            
        except Exception:
            return False
    
    def _is_date_string(self, text):
        """Check if text looks like a date in MM/DD/YYYY format"""
        try:
            if '/' in text and len(text.split('/')) == 3:
                parts = text.split('/')
                return (len(parts[0]) <= 2 and len(parts[1]) <= 2 and 
                       len(parts[2]) == 4 and all(p.isdigit() for p in parts))
            return False
        except:
            return False
    
    def extract_individuals_from_excel(self, workbook_path):
        """
        Extract individuals (DD, DM, OT, etc.) from Excel file header rows
        """
        try:
            wb = openpyxl.load_workbook(workbook_path)
            ws = wb['DailyMatrix'] if 'DailyMatrix' in wb.sheetnames else wb.active
            
            individuals = []
            
            # Look for header row with individuals
            for row_num in range(1, min(20, ws.max_row + 1)):
                cell_A = ws.cell(row=row_num, column=1)
                if cell_A.value and 'Service Provider' in str(cell_A.value):
                    # Found header row, extract individuals from columns
                    for col_num in range(2, ws.max_column + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value and str(cell.value).strip():
                            value = str(cell.value).strip()
                            # Skip 'Provider Total' column
                            if value not in ['Provider Total', 'Total']:
                                individuals.append(value)
                    break
            
            # Default to DD, DM, OT if not found
            if not individuals:
                individuals = ['DD', 'DM', 'OT']
                st.info("Using default individuals: DD, DM, OT")
            else:
                st.success(f"Extracted individuals from file: {', '.join(individuals)}")
            
            self.expected_individuals = individuals
            return individuals
            
        except Exception as e:
            st.error(f"Error extracting individuals: {str(e)}")
            self.expected_individuals = ['DD', 'DM', 'OT']
            return self.expected_individuals
    
    def extract_day_blocks_with_formatting(self, workbook_path):
        """
        Extract day blocks with original formatting preserved
        Returns structured data with formatting information
        """
        try:
            # First extract individuals from the file
            self.extract_individuals_from_excel(workbook_path)
            
            wb = openpyxl.load_workbook(workbook_path)
            ws = wb['DailyMatrix'] if 'DailyMatrix' in wb.sheetnames else wb.active
            
            day_blocks = []
            current_day = None
            day_start_row = None
            
            for row_num in range(1, ws.max_row + 1):
                cell_A = ws.cell(row=row_num, column=1)
                cell_value = str(cell_A.value).strip() if cell_A.value else ""
                
                # Check if this is a date row
                if self._is_date_string(cell_value):
                    # Save previous day if exists
                    if current_day and day_start_row:
                        day_block = self._extract_day_block(ws, day_start_row, row_num - 1, current_day)
                        if day_block:
                            day_blocks.append(day_block)
                    
                    # Start new day
                    current_day = cell_value
                    day_start_row = row_num
            
            # Don't forget the last day
            if current_day and day_start_row:
                day_block = self._extract_day_block(ws, day_start_row, ws.max_row, current_day)
                if day_block:
                    day_blocks.append(day_block)
            
            return day_blocks, wb
            
        except Exception as e:
            st.error(f"Error extracting day blocks with formatting: {str(e)}")
            return [], None
    
    def _extract_day_block(self, worksheet, start_row, end_row, date_str):
        """
        Extract a single day block with formatting information
        """
        day_data = {
            'date': date_str,
            'date_row': start_row,
            'providers': [],
            'totals_row': None,
            'pending_row': None,
            'formatting': {},
            'formulas': {}
        }
        
        # Scan through the day block
        for row_num in range(start_row, end_row + 1):
            cell_A = worksheet.cell(row=row_num, column=1)
            cell_value = str(cell_A.value).strip() if cell_A.value else ""
            
            # Skip empty rows and the date row itself
            if not cell_value or cell_value == date_str:
                continue
            
            # Check for header row
            if 'Service Provider' in cell_value:
                continue
            
            # Check for totals row
            elif 'Total hours for individual' in cell_value:
                day_data['totals_row'] = row_num
                day_data['formulas'][row_num] = self._extract_row_formulas(worksheet, row_num)
            
            # Check for pending row
            elif 'Total hrs pending' in cell_value:
                day_data['pending_row'] = row_num
                day_data['formulas'][row_num] = self._extract_row_formulas(worksheet, row_num)
            
            # This must be a provider row
            else:
                provider_data = self._extract_provider_row(worksheet, row_num, cell_value)
                if provider_data:
                    day_data['providers'].append(provider_data)
        
        # Extract formatting information for the entire block
        day_data['formatting'] = self._extract_block_formatting(worksheet, start_row, end_row)
        
        return day_data
    
    def _extract_provider_row(self, worksheet, row_num, provider_name):
        """
        Extract provider row data with formatting
        """
        provider_data = {
            'name': provider_name,
            'row': row_num,
            'hours': {},
            'total_formula': None,
            'formatting': {}
        }
        
        # Extract hours for individuals (columns B, C, D, etc.)
        for col_idx, individual in enumerate(self.expected_individuals, 2):  # Start from column B (2)
            cell = worksheet.cell(row=row_num, column=col_idx)
            provider_data['hours'][individual] = self._extract_numeric_value(cell.value)
            
            # Store formatting info
            provider_data['formatting'][individual] = {
                'fill': cell.fill.start_color.index if cell.fill.start_color.index != '00000000' else None,
                'font_color': self._get_font_color(cell.font),
                'bold': cell.font.bold
            }
        
        # Extract total formula (column E)
        total_cell = worksheet.cell(row=row_num, column=5)
        if total_cell.value and str(total_cell.value).startswith('='):
            provider_data['total_formula'] = str(total_cell.value)
        else:
            provider_data['total_formula'] = f"=SUM(B{row_num}:D{row_num})"
        
        # Check if this is a newly added provider (green font)
        name_cell = worksheet.cell(row=row_num, column=1)
        provider_data['is_new_provider'] = (
            self._get_font_color(name_cell.font) == self.colors['GREEN_FONT']
        )
        
        return provider_data
    
    def _extract_row_formulas(self, worksheet, row_num):
        """
        Extract formulas from a row
        """
        formulas = {}
        for col_num in range(1, 6):  # A through E
            cell = worksheet.cell(row=row_num, column=col_num)
            if cell.value and str(cell.value).startswith('='):
                formulas[col_num] = str(cell.value)
        return formulas
    
    def _extract_block_formatting(self, worksheet, start_row, end_row):
        """
        Extract formatting information for an entire day block
        """
        formatting = {}
        
        for row_num in range(start_row, end_row + 1):
            row_formatting = {}
            for col_num in range(1, 6):  # A through E
                cell = worksheet.cell(row=row_num, column=col_num)
                row_formatting[col_num] = {
                    'fill': cell.fill.start_color.index if cell.fill.start_color.index != '00000000' else None,
                    'font_color': self._get_font_color(cell.font),
                    'bold': cell.font.bold,
                    'value': cell.value
                }
            formatting[row_num] = row_formatting
        
        return formatting
    
    def _get_font_color(self, font):
        """
        Extract font color from font object
        """
        try:
            if font.color and hasattr(font.color, 'rgb') and font.color.rgb:
                return font.color.rgb
        except:
            pass
        return None
    
    def _extract_numeric_value(self, cell_value):
        """
        Extract numeric value from cell, handling formulas and text
        """
        try:
            if cell_value is None:
                return 0
            
            cell_str = str(cell_value).strip()
            
            if cell_str in ['nan', '', 'None']:
                return 0
            
            # Handle Excel formulas (return 0 for now, will be calculated)
            if cell_str.startswith('='):
                return 0
            
            return float(cell_str)
            
        except:
            return 0
    
    def validate_day_blocks(self, day_blocks):
        """
        Validate all day blocks for consistency
        """
        all_issues = []
        
        for day_data in day_blocks:
            issues = self._validate_single_day(day_data)
            all_issues.extend(issues)
        
        return all_issues
    
    def _validate_single_day(self, day_data):
        """
        Validate a single day's data
        """
        issues = []
        
        if not day_data['providers']:
            issues.append(f"No providers found for {day_data['date']}")
        
        # Additional validation can be added here
        
        return issues
