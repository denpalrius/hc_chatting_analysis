"""
Excel Parser Module

Parses Excel files containing schedule data and extracts structured day blocks
with provider information for processing by the schedule balancer.
"""

import logging
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re
from datetime import datetime


class ExcelParser:
    """
    Parses Excel files containing schedule data into structured day blocks.
    
    The parser identifies date headers and extracts provider information with
    hours allocated to different individuals (DD, DM, OT, etc.).
    """
    
    def __init__(self, individuals: List[str] = None):
        """
        Initialize the Excel parser.
        
        Args:
            individuals: List of individual types to look for (e.g., ['DD', 'DM', 'OT'])
        """
        self.individuals = individuals or ['DD', 'DM', 'OT']
        self.logger = logging.getLogger(__name__)
        
    def parse_excel_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Parse an Excel file and extract day blocks with provider data.
        
        Args:
            file_path: Path to the Excel file to parse
            
        Returns:
            List of day blocks, where each day block contains:
            - date: The date string
            - providers: List of provider dictionaries with name and hours
            - row_info: Additional row information for tracking
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            day_blocks = self._extract_day_blocks(worksheet)
            
            self.logger.info(f"Parsed {len(day_blocks)} day blocks from {file_path}")
            return day_blocks
            
        except Exception as e:
            self.logger.error(f"Error parsing Excel file {file_path}: {str(e)}")
            return []
    
    def _extract_day_blocks(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Extract day blocks from the worksheet.
        
        Args:
            worksheet: The Excel worksheet to parse
            
        Returns:
            List of day block dictionaries
        """
        day_blocks = []
        current_day_block = None
        
        # Find header row to determine column mapping
        header_mapping = self._find_header_mapping(worksheet)
        if not header_mapping:
            self.logger.warning("Could not find header mapping in worksheet")
            return []
        
        # Process each row
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            row_data = [cell.value for cell in row]
            
            # Skip empty rows
            if not any(row_data):
                continue
                
            # Check if this is a date row
            date_str = self._extract_date_from_row(row_data)
            if date_str:
                # Save previous day block if exists
                if current_day_block:
                    day_blocks.append(current_day_block)
                
                # Start new day block
                current_day_block = {
                    'date': date_str,
                    'providers': [],
                    'row_info': {
                        'date_row': row_idx,
                        'provider_rows': []
                    }
                }
                continue
            
            # Check if this is a provider row
            if current_day_block and self._is_provider_row(row_data, header_mapping):
                provider = self._extract_provider_data(row_data, header_mapping, row_idx)
                if provider:
                    current_day_block['providers'].append(provider)
                    current_day_block['row_info']['provider_rows'].append(row_idx)
                continue
            
            # Check if this is a totals row
            if current_day_block and self._is_totals_row(row_data):
                current_day_block['row_info']['totals_row'] = row_idx
                continue
            
            # Check if this is a pending row
            if current_day_block and self._is_pending_row(row_data):
                current_day_block['row_info']['pending_row'] = row_idx
                continue
        
        # Add the last day block
        if current_day_block:
            day_blocks.append(current_day_block)
            
        return day_blocks
    
    def _find_header_mapping(self, worksheet: Worksheet) -> Optional[Dict[str, int]]:
        """
        Find the header row and create column mapping.
        
        Args:
            worksheet: The Excel worksheet
            
        Returns:
            Dictionary mapping column names to column indices (0-based)
        """
        for row in worksheet.iter_rows(max_row=10):  # Check first 10 rows
            row_data = [str(cell.value).strip().upper() if cell.value else '' for cell in row]
            
            mapping = {}
            
            # Look for provider name column
            for idx, value in enumerate(row_data):
                if 'PROVIDER' in value or 'NAME' in value:
                    mapping['provider_name'] = idx
                    break
            
            if 'provider_name' not in mapping:
                continue
            
            # Look for individual columns
            for individual in self.individuals:
                for idx, value in enumerate(row_data):
                    if value == individual.upper():
                        mapping[individual] = idx
                        break
            
            # Look for total hours column
            for idx, value in enumerate(row_data):
                if 'TOTAL' in value and 'HOUR' in value:
                    mapping['total_hours'] = idx
                    break
            
            # If we found provider name and at least one individual, we have a valid header
            if 'provider_name' in mapping and any(ind in mapping for ind in self.individuals):
                return mapping
        
        return None
    
    def _extract_date_from_row(self, row_data: List[Any]) -> Optional[str]:
        """
        Extract date string from a row if it contains a date.
        
        Args:
            row_data: List of cell values from the row
            
        Returns:
            Date string if found, None otherwise
        """
        for value in row_data:
            if not value:
                continue
                
            value_str = str(value).strip()
            
            # Check for date patterns
            date_patterns = [
                r'\d{1,2}/\d{1,2}/\d{4}',  # MM/DD/YYYY or M/D/YYYY
                r'\d{4}-\d{1,2}-\d{1,2}',  # YYYY-MM-DD
                r'\d{1,2}-\d{1,2}-\d{4}',  # MM-DD-YYYY
            ]
            
            for pattern in date_patterns:
                if re.search(pattern, value_str):
                    return value_str
            
            # Check if it's a datetime object
            if isinstance(value, datetime):
                return value.strftime('%m/%d/%Y')
        
        return None
    
    def _is_provider_row(self, row_data: List[Any], header_mapping: Dict[str, int]) -> bool:
        """
        Check if a row contains provider data.
        
        Args:
            row_data: List of cell values from the row
            header_mapping: Column mapping dictionary
            
        Returns:
            True if this is a provider row, False otherwise
        """
        if 'provider_name' not in header_mapping:
            return False
            
        provider_name_idx = header_mapping['provider_name']
        
        # Check if there's a provider name
        if provider_name_idx >= len(row_data):
            return False
            
        provider_name = row_data[provider_name_idx]
        if not provider_name:
            return False
        
        provider_name_str = str(provider_name).strip()
        
        # Skip empty names, totals, and obvious non-provider entries
        if not provider_name_str or provider_name_str.upper() in ['TOTALS', 'TOTAL', '']:
            return False
        
        # Must contain some alphabetic characters (provider names have letters)
        if not re.search(r'[A-Za-z]', provider_name_str):
            return False
            
        return True
    
    def _is_totals_row(self, row_data: List[Any]) -> bool:
        """
        Check if a row is a totals row.
        
        Args:
            row_data: List of cell values from the row
            
        Returns:
            True if this is a totals row, False otherwise
        """
        if not row_data:
            return False
            
        first_value = str(row_data[0]).strip().upper() if row_data[0] else ''
        return first_value in ['TOTALS', 'TOTAL']
    
    def _is_pending_row(self, row_data: List[Any]) -> bool:
        """
        Check if a row is a pending row.
        
        Args:
            row_data: List of cell values from the row
            
        Returns:
            True if this is a pending row, False otherwise
        """
        if not row_data:
            return False
            
        first_value = str(row_data[0]).strip().upper() if row_data[0] else ''
        return first_value in ['PENDING', 'PEND']
    
    def _extract_provider_data(self, row_data: List[Any], header_mapping: Dict[str, int], row_idx: int) -> Optional[Dict[str, Any]]:
        """
        Extract provider data from a row.
        
        Args:
            row_data: List of cell values from the row
            header_mapping: Column mapping dictionary
            row_idx: Row index in the worksheet
            
        Returns:
            Provider dictionary or None if extraction fails
        """
        try:
            provider_name_idx = header_mapping['provider_name']
            provider_name = str(row_data[provider_name_idx]).strip()
            
            if not provider_name:
                return None
            
            # Extract hours for each individual
            hours = {}
            for individual in self.individuals:
                if individual in header_mapping:
                    col_idx = header_mapping[individual]
                    if col_idx < len(row_data):
                        value = row_data[col_idx]
                        try:
                            # Convert to float, default to 0 if not a number
                            hours[individual] = float(value) if value is not None else 0.0
                        except (ValueError, TypeError):
                            hours[individual] = 0.0
                    else:
                        hours[individual] = 0.0
                else:
                    hours[individual] = 0.0
            
            return {
                'name': provider_name,
                'hours': hours,
                'row': row_idx
            }
            
        except Exception as e:
            self.logger.warning(f"Error extracting provider data from row {row_idx}: {str(e)}")
            return None
    
    def get_column_mapping(self, worksheet: Worksheet) -> Optional[Dict[str, int]]:
        """
        Get the column mapping for the worksheet.
        
        Args:
            worksheet: The Excel worksheet
            
        Returns:
            Dictionary mapping column names to column indices
        """
        return self._find_header_mapping(worksheet)
