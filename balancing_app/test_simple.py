#!/usr/bin/env python3
"""
Simple test script for core functionality without streamlit
"""

import sys
import os
import pandas as pd
import openpyxl

def test_basic_functionality():
    """Test basic file reading and processing"""
    print("🧪 Testing Basic Functionality")
    print("=" * 40)
    
    # Test file path
    sample_file = "../data/july_summary.xlsx"
    
    if not os.path.exists(sample_file):
        print(f"❌ Sample file not found: {sample_file}")
        return False
    
    try:
        # Test pandas reading
        print("📖 Testing pandas file reading...")
        df = pd.read_excel(sample_file, engine='openpyxl')
        print(f"✅ Read Excel file: {df.shape}")
        
        # Test openpyxl reading
        print("📖 Testing openpyxl file reading...")
        wb = openpyxl.load_workbook(sample_file)
        ws = wb.active
        print(f"✅ Loaded workbook with {len(wb.sheetnames)} sheets")
        print(f"   Active sheet: {ws.title}, Size: {ws.max_row}x{ws.max_column}")
        
        # Test date extraction
        print("📅 Testing date extraction...")
        dates_found = []
        for row_num in range(1, min(50, ws.max_row + 1)):
            cell = ws.cell(row=row_num, column=1)
            if cell.value and '/' in str(cell.value):
                if len(str(cell.value).split('/')) == 3:
                    dates_found.append((row_num, str(cell.value)))
        
        print(f"✅ Found {len(dates_found)} date entries:")
        for i, (row, date) in enumerate(dates_found[:5]):
            print(f"   Row {row}: {date}")
        
        # Test provider extraction
        print("👥 Testing provider extraction...")
        providers_found = []
        for row_num in range(1, min(50, ws.max_row + 1)):
            cell = ws.cell(row=row_num, column=1)
            if cell.value and 'RN' in str(cell.value):
                providers_found.append((row_num, str(cell.value)))
        
        print(f"✅ Found {len(providers_found)} provider entries:")
        for i, (row, provider) in enumerate(providers_found[:3]):
            print(f"   Row {row}: {provider}")
        
        print("🎉 Basic functionality test passed!")
        return True
        
    except Exception as e:
        print(f"❌ Test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main test function"""
    print("Healthcare Schedule Balancer - Basic Test")
    print("=" * 50)
    
    success = test_basic_functionality()
    
    if success:
        print("\n✅ Basic tests completed successfully!")
        print("🚀 Ready to run the Streamlit app: ./run.sh")
    else:
        print("\n❌ Basic tests failed.")
        sys.exit(1)

if __name__ == "__main__":
    main()
