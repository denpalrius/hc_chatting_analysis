import streamlit as st
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
from copy import deepcopy

class ScheduleBalancer:
    """
    Implements healthcare provider schedule balancing according to business rules
    """
    
    def __init__(self, individuals=None, additional_providers=None):
        self.changes_log = []
        self.individuals = individuals or ['DD', 'DM', 'OT']  # Will be set dynamically
        
        # Core providers and supplemental providers
        self.supplemental_providers = [
            'Charles Sagini, RN/House Manager',
            'Josephine Sagini, RN/Program Manager', 
            'Faith Murerwa, RN/House Supervisor'
        ]
        self.additional_providers = additional_providers or []
        
        # Color definitions
        self.colors = {
            'RED_FILL': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),
            'YELLOW_FILL': PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid'),
            'GREEN_FILL': PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid'),
            'ORANGE_FILL': PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid'),
            'GREEN_FONT': Font(color='FF00B050')
        }
    
    def balance_schedule(self, day_blocks, wb):
        """
        Balance entire schedule according to business rules
        """
        summary = {
            'total_days_processed': len(day_blocks),
            'days_balanced': 0,
            'days_unbalanced': 0,
            'providers_added': 0,
            'entries_modified': 0
        }
        
        ws = wb.active if wb else None
        if not ws:
            st.error("No active worksheet found")
            return None, summary
        
        for day_data in day_blocks:
            try:
                balanced, modifications = self._balance_single_day(day_data, ws)
                
                if balanced:
                    summary['days_balanced'] += 1
                else:
                    summary['days_unbalanced'] += 1
                    # Mark date cell as red for unbalanced days
                    self._mark_unbalanced_date(ws, day_data['date_row'])
                
                summary['entries_modified'] += modifications['entries_modified']
                summary['providers_added'] += modifications['providers_added']
                
                self.changes_log.extend(modifications['log'])
                
            except Exception as e:
                st.error(f"Error balancing {day_data['date']}: {str(e)}")
                summary['days_unbalanced'] += 1
        
        return wb, summary
    
    def _balance_single_day(self, day_data, worksheet):
        """
        Balance a single day according to business rules priority
        """
        modifications = {
            'entries_modified': 0,
            'providers_added': 0,
            'log': []
        }
        
        date = day_data['date']
        
        # Calculate current totals for each individual
        current_totals = {individual: 0 for individual in self.individuals}
        for provider in day_data['providers']:
            for individual in self.individuals:
                if individual in provider['hours']:
                    current_totals[individual] += provider['hours'][individual]
        
        modifications['log'].append(f"{date}: Starting totals - {current_totals}")
        
        # STEP 1: Identify and fix over-allocated providers (>16 hours)
        modifications['log'].append(f"{date}: Step 1 - Fixing over-allocated providers")
        self._fix_over_allocated_providers(day_data, worksheet, modifications)
        
        # STEP 2: Recalculate totals after fixing over-allocations
        current_totals = {individual: 0 for individual in self.individuals}
        for provider in day_data['providers']:
            for individual in self.individuals:
                if individual in provider['hours']:
                    current_totals[individual] += provider['hours'][individual]
        
        modifications['log'].append(f"{date}: After over-allocation fix - {current_totals}")
        
        # STEP 3: Calculate pending hours and balance each individual
        pending_hours = {individual: 24 - current_totals[individual] for individual in self.individuals}
        modifications['log'].append(f"{date}: Pending hours - {pending_hours}")
        
        # Check if already balanced
        if all(pending == 0 for pending in pending_hours.values()):
            modifications['log'].append(f"{date}: Already balanced after over-allocation fix")
            self._update_totals_formulas(day_data, worksheet)
            return True, modifications
        
        # STEP 4: Balance each individual by adding supplemental providers
        modifications['log'].append(f"{date}: Step 2 - Adding supplemental providers to fill gaps")
        balanced = True
        
        for individual in self.individuals:
            if pending_hours[individual] > 0:  # Need more hours
                success = self._add_hours_for_individual(day_data, worksheet, individual, 
                                                        pending_hours[individual], modifications)
                if not success:
                    balanced = False
                    modifications['log'].append(f"{date}: Could not balance {individual} (still need {pending_hours[individual]} hours)")
        
        # Final verification
        final_totals = {individual: 0 for individual in self.individuals}
        for provider in day_data['providers']:
            for individual in self.individuals:
                if individual in provider['hours']:
                    final_totals[individual] += provider['hours'][individual]
        
        final_pending = {individual: 24 - final_totals[individual] for individual in self.individuals}
        balanced = all(pending == 0 for pending in final_pending.values())
        
        if balanced:
            modifications['log'].append(f"{date}: ✅ Successfully balanced - {final_totals}")
        else:
            modifications['log'].append(f"{date}: ❌ Failed to balance - Final totals: {final_totals}, Pending: {final_pending}")
        
        # Update totals and pending formulas
        self._update_totals_formulas(day_data, worksheet)
        
        return balanced, modifications
    
    def _balance_individual(self, day_data, worksheet, individual, pending_hours, modifications):
        """
        Balance hours for a specific individual (DD, DM, or OT)
        """
        if pending_hours == 0:
            return True
        
        date = day_data['date']
        
        # Rule 1: Try to adjust existing providers (max 16 hours, min 2 hours)
        if self._try_adjust_existing_providers(day_data, worksheet, individual, pending_hours, modifications):
            return True
        
        # Rule 2: Add supplemental providers
        if self._try_add_supplemental_providers(day_data, worksheet, individual, pending_hours, modifications):
            return True
        
        # Exception Rule 1: Modify non-zero entries for existing providers
        if self._try_modify_existing_nonzero(day_data, worksheet, individual, pending_hours, modifications):
            return True
        
        # Exception Rule 2: Raise cap to 18 hours
        if self._try_raise_cap_to_18(day_data, worksheet, individual, pending_hours, modifications):
            return True
        
        # Exception Rule 3: Add Carolyn Porter for OT only
        if individual == 'OT' and self._try_add_carolyn_porter(day_data, worksheet, individual, pending_hours, modifications):
            return True
        
        # If all rules fail, mark as unbalanced
        return False
    
    def _try_adjust_existing_providers(self, day_data, worksheet, individual, pending_hours, modifications):
        """
        Try to adjust existing providers within normal limits (2-16 hours)
        """
        providers = day_data['providers']
        
        for provider in providers:
            current_hours = provider['hours'][individual]
            current_total = sum(provider['hours'].values())
            
            if pending_hours > 0:  # Need to add hours
                max_additional = min(16 - current_total, pending_hours)
                if max_additional > 0 and current_hours + max_additional >= 2:
                    # Add hours
                    new_hours = current_hours + max_additional
                    self._update_provider_hours(worksheet, provider, individual, new_hours, 'GREEN_FILL')
                    modifications['entries_modified'] += 1
                    modifications['log'].append(f"{day_data['date']}: Increased {provider['name']} {individual} from {current_hours} to {new_hours}")
                    pending_hours -= max_additional
                    if pending_hours == 0:
                        return True
            
            elif pending_hours < 0:  # Need to reduce hours
                max_reduction = min(current_hours - 2, -pending_hours) if current_hours > 2 else 0
                if max_reduction > 0:
                    # Reduce hours
                    new_hours = current_hours - max_reduction
                    self._update_provider_hours(worksheet, provider, individual, new_hours, 'ORANGE_FILL')
                    modifications['entries_modified'] += 1
                    modifications['log'].append(f"{day_data['date']}: Reduced {provider['name']} {individual} from {current_hours} to {new_hours}")
                    pending_hours += max_reduction
                    if pending_hours == 0:
                        return True
        
        return pending_hours == 0
    
    def _try_add_supplemental_providers(self, day_data, worksheet, individual, pending_hours, modifications):
        """
        Try to add supplemental providers or modify existing zero entries
        """
        if pending_hours <= 0:
            return True
        
        # First, try to modify existing supplemental providers with 0 hours
        for provider in day_data['providers']:
            if (provider['name'] in self.supplemental_providers and 
                provider['hours'][individual] == 0):
                
                hours_to_add = min(16, pending_hours)
                self._update_provider_hours(worksheet, provider, individual, hours_to_add, 'GREEN_FILL')
                modifications['entries_modified'] += 1
                modifications['log'].append(f"{day_data['date']}: Set {provider['name']} {individual} to {hours_to_add} hours")
                pending_hours -= hours_to_add
                
                if pending_hours == 0:
                    return True
        
        # If still need hours, add new supplemental providers
        for supp_provider in self.supplemental_providers:
            if pending_hours <= 0:
                break
                
            # Check if this provider is already in the day
            existing_provider = next((p for p in day_data['providers'] if supp_provider in p['name']), None)
            if existing_provider:
                continue  # Already handled above
            
            # Add new provider
            hours_to_add = min(16, pending_hours)
            new_row = self._add_new_provider_row(day_data, worksheet, supp_provider, individual, hours_to_add)
            if new_row:
                modifications['providers_added'] += 1
                modifications['entries_modified'] += 1
                modifications['log'].append(f"{day_data['date']}: Added {supp_provider} with {hours_to_add} {individual} hours")
                pending_hours -= hours_to_add
        
        return pending_hours == 0
    
    def _try_modify_existing_nonzero(self, day_data, worksheet, individual, pending_hours, modifications):
        """
        Exception Rule 1: Modify existing non-zero entries
        """
        for provider in day_data['providers']:
            current_hours = provider['hours'][individual]
            current_total = sum(provider['hours'].values())
            
            if current_hours > 0:  # Non-zero entry
                if pending_hours > 0:  # Need more hours
                    max_additional = 16 - current_total
                    if max_additional > 0:
                        hours_to_add = min(max_additional, pending_hours)
                        new_hours = current_hours + hours_to_add
                        self._update_provider_hours(worksheet, provider, individual, new_hours, 'ORANGE_FILL')
                        modifications['entries_modified'] += 1
                        modifications['log'].append(f"{day_data['date']}: Modified {provider['name']} {individual} from {current_hours} to {new_hours}")
                        pending_hours -= hours_to_add
                        
                        if pending_hours == 0:
                            return True
        
        return pending_hours == 0
    
    def _try_raise_cap_to_18(self, day_data, worksheet, individual, pending_hours, modifications):
        """
        Exception Rule 2: Temporarily raise provider cap to 18 hours
        """
        for provider in day_data['providers']:
            current_total = sum(provider['hours'].values())
            current_hours = provider['hours'][individual]
            
            if current_total < 18:  # Can still add more hours
                max_additional = 18 - current_total
                hours_to_add = min(max_additional, pending_hours)
                
                if hours_to_add > 0:
                    new_hours = current_hours + hours_to_add
                    self._update_provider_hours(worksheet, provider, individual, new_hours, 'ORANGE_FILL')
                    modifications['entries_modified'] += 1
                    modifications['log'].append(f"{day_data['date']}: Raised cap - Modified {provider['name']} {individual} from {current_hours} to {new_hours}")
                    pending_hours -= hours_to_add
                    
                    if pending_hours == 0:
                        return True
        
        return pending_hours == 0
    
    def _try_add_carolyn_porter(self, day_data, worksheet, individual, pending_hours, modifications):
        """
        Exception Rule 3: Add Carolyn Porter, LPN (only for OT)
        """
        if individual != 'OT' or pending_hours <= 0:
            return False
        
        # Check if Carolyn is already in the day
        existing_carolyn = next((p for p in day_data['providers'] if 'Carolyn' in p['name']), None)
        
        if existing_carolyn:
            # Modify existing Carolyn entry
            current_hours = existing_carolyn['hours']['OT']
            current_total = sum(existing_carolyn['hours'].values())
            max_additional = min(16 - current_total, pending_hours)
            
            if max_additional > 0:
                new_hours = current_hours + max_additional
                self._update_provider_hours(worksheet, existing_carolyn, 'OT', new_hours, 'GREEN_FILL')
                modifications['entries_modified'] += 1
                modifications['log'].append(f"{day_data['date']}: Modified Carolyn Porter OT from {current_hours} to {new_hours}")
                return True
        else:
            # Add new Carolyn Porter row
            hours_to_add = min(16, pending_hours)
            new_row = self._add_new_provider_row(day_data, worksheet, self.carolyn_porter, 'OT', hours_to_add)
            if new_row:
                modifications['providers_added'] += 1
                modifications['entries_modified'] += 1
                modifications['log'].append(f"{day_data['date']}: Added Carolyn Porter with {hours_to_add} OT hours")
                return True
        
        return False
    
    def _update_provider_hours(self, worksheet, provider, individual, new_hours, color_type):
        """
        Update provider hours in worksheet with appropriate formatting
        """
        # Map individual to column dynamically
        col_map = {ind: idx + 2 for idx, ind in enumerate(self.individuals)}  # Start from column B (2)
        col_num = col_map.get(individual, 2)  # Default to column B if not found
        row_num = provider['row']
        
        # Update the cell value
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = new_hours
        
        # Apply formatting
        if color_type in self.colors:
            if 'FILL' in color_type:
                cell.fill = self.colors[color_type]
            elif 'FONT' in color_type:
                cell.font = self.colors[color_type]
        
        # Update provider data
        provider['hours'][individual] = new_hours
    
    def _add_new_provider_row(self, day_data, worksheet, provider_name, individual, hours):
        """
        Add a new provider row to the day block
        """
        # Find where to insert the new row (before totals row)
        totals_row = day_data.get('totals_row')
        if not totals_row:
            return None
        
        insert_row = totals_row
        
        # Insert new row
        worksheet.insert_rows(insert_row)
        
        # Update all row references in day_data
        self._update_row_references(day_data, insert_row)
        
        # Set up the new provider row
        col_map = {'DD': 2, 'DM': 3, 'OT': 4}  # B, C, D columns
        
        # Provider name (column A) with green font for new providers
        name_cell = worksheet.cell(row=insert_row, column=1)
        name_cell.value = provider_name
        name_cell.font = self.colors['GREEN_FONT']
        
        # Hours for each individual (columns B, C, D)
        for ind in self.individuals:
            cell = worksheet.cell(row=insert_row, column=col_map[ind])
            cell.value = hours if ind == individual else 0
            if ind == individual and hours > 0:
                cell.fill = self.colors['GREEN_FILL']
        
        # Total formula (column E)
        total_cell = worksheet.cell(row=insert_row, column=5)
        total_cell.value = f"=SUM(B{insert_row}:D{insert_row})"
        
        # Add to day_data providers list
        new_provider = {
            'name': provider_name,
            'row': insert_row,
            'hours': {ind: hours if ind == individual else 0 for ind in self.individuals},
            'total_formula': f"=SUM(B{insert_row}:D{insert_row})",
            'is_new_provider': True
        }
        day_data['providers'].append(new_provider)
        
        return insert_row
    
    def _update_row_references(self, day_data, inserted_row):
        """
        Update row references after inserting a new row
        """
        # Update totals and pending row numbers
        if day_data.get('totals_row') and day_data['totals_row'] >= inserted_row:
            day_data['totals_row'] += 1
        
        if day_data.get('pending_row') and day_data['pending_row'] >= inserted_row:
            day_data['pending_row'] += 1
        
        # Update provider row references
        for provider in day_data['providers']:
            if provider['row'] >= inserted_row:
                provider['row'] += 1
                # Update total formula
                new_row = provider['row']
                provider['total_formula'] = f"=SUM(B{new_row}:D{new_row})"
    
    def _update_totals_formulas(self, day_data, worksheet):
        """
        Update the totals and pending hours formulas
        """
        totals_row = day_data.get('totals_row')
        pending_row = day_data.get('pending_row')
        
        if totals_row:
            # Update totals formulas to include all provider rows
            provider_rows = [p['row'] for p in day_data['providers']]
            if provider_rows:
                min_row = min(provider_rows)
                max_row = max(provider_rows)
                
                # Update totals for DD, DM, OT (columns B, C, D)
                for col_num in range(2, 5):
                    cell = worksheet.cell(row=totals_row, column=col_num)
                    cell.value = f"=SUM({chr(64 + col_num)}{min_row}:{chr(64 + col_num)}{max_row})"
        
        if pending_row:
            # Update pending formulas
            for col_num in range(2, 5):  # B, C, D columns
                cell = worksheet.cell(row=pending_row, column=col_num)
                cell.value = f"=24-{chr(64 + col_num)}{totals_row}" if totals_row else "=24"
    
    def _fix_over_allocated_providers(self, day_data, worksheet, modifications):
        """
        Fix providers who exceed 16 hours by reducing their allocation
        """
        date = day_data['date']
        
        for provider in day_data['providers']:
            total_hours = sum(provider['hours'].values())
            
            if total_hours > 16:
                modifications['log'].append(f"{date}: Provider {provider['name']} over-allocated with {total_hours} hours")
                
                # Reduce hours proportionally or by priority
                excess = total_hours - 16
                
                # Priority: OT first, then DM, then DD (as per business rules)
                reduction_priority = ['OT', 'DM', 'DD']
                
                for individual in reduction_priority:
                    if individual in self.individuals and excess > 0:
                        current_hours = provider['hours'][individual]
                        if current_hours > 0:
                            # Reduce as much as possible while maintaining minimum of 2 hours
                            min_hours = 2 if current_hours >= 2 else 0
                            max_reduction = min(current_hours - min_hours, excess)
                            
                            if max_reduction > 0:
                                new_hours = current_hours - max_reduction
                                self._update_provider_hours(worksheet, provider, individual, new_hours, 'ORANGE_FILL')
                                modifications['entries_modified'] += 1
                                modifications['log'].append(f"{date}: Reduced {provider['name']} {individual} from {current_hours} to {new_hours} (over-allocation fix)")
                                excess -= max_reduction
                
                if excess > 0:
                    modifications['log'].append(f"{date}: ⚠️ Could not fully fix over-allocation for {provider['name']}, remaining excess: {excess}")
    
    def _add_hours_for_individual(self, day_data, worksheet, individual, pending_hours, modifications):
        """
        Add hours for a specific individual by adding supplemental providers
        """
        if pending_hours <= 0:
            return True
        
        date = day_data['date']
        modifications['log'].append(f"{date}: Need to add {pending_hours} hours for {individual}")
        
        # First, try to modify existing supplemental providers with 0 hours
        for provider in day_data['providers']:
            if (provider['name'] in self.supplemental_providers and 
                provider['hours'][individual] == 0):
                
                hours_to_add = min(16, pending_hours)
                self._update_provider_hours(worksheet, provider, individual, hours_to_add, 'GREEN_FILL')
                modifications['entries_modified'] += 1
                modifications['log'].append(f"{date}: Set {provider['name']} {individual} to {hours_to_add} hours")
                pending_hours -= hours_to_add
                
                if pending_hours == 0:
                    return True
        
        # If still need hours, add new supplemental providers
        for supp_provider in self.supplemental_providers:
            if pending_hours <= 0:
                break
                
            # Check if this provider is already in the day
            existing_provider = next((p for p in day_data['providers'] if supp_provider in p['name']), None)
            if existing_provider:
                continue  # Already handled above
            
            # Add new provider
            hours_to_add = min(16, pending_hours)
            new_row = self._add_new_provider_row(day_data, worksheet, supp_provider, individual, hours_to_add)
            if new_row:
                modifications['providers_added'] += 1
                modifications['entries_modified'] += 1
                modifications['log'].append(f"{date}: Added {supp_provider} with {hours_to_add} {individual} hours")
                pending_hours -= hours_to_add
        
        # Final check - if still need hours, try to increase existing providers within limits
        if pending_hours > 0:
            for provider in day_data['providers']:
                current_total = sum(provider['hours'].values())
                if current_total < 16:  # Can still add hours
                    max_additional = min(16 - current_total, pending_hours)
                    if max_additional > 0:
                        current_hours = provider['hours'][individual]
                        new_hours = current_hours + max_additional
                        self._update_provider_hours(worksheet, provider, individual, new_hours, 'YELLOW_FILL')
                        modifications['entries_modified'] += 1
                        modifications['log'].append(f"{date}: Increased {provider['name']} {individual} from {current_hours} to {new_hours}")
                        pending_hours -= max_additional
                        
                        if pending_hours == 0:
                            return True
        
        return pending_hours == 0
    
    def _mark_unbalanced_date(self, worksheet, date_row):
        """
        Mark the date cell as red for unbalanced days
        """
        date_cell = worksheet.cell(row=date_row, column=1)
        date_cell.fill = self.colors['RED_FILL']
    
    def get_changes_log(self):
        """
        Get the complete log of changes made
        """
        return self.changes_log
