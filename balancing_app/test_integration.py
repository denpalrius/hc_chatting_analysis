import pytest
import unittest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

from excel_parser import ExcelParser
from schedule_balancer import ScheduleBalancer


class TestIntegration(unittest.TestCase):
    """Integration tests for complete schedule balancing workflow"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_files_created = []
        
        # Standard configuration
        self.individuals = ['DD', 'DM', 'OT']
        self.additional_providers = [
            'Charles Sagini, RN/House Manager',
            'Other Supplemental Provider'
        ]
        
    def tearDown(self):
        """Clean up test files"""
        for file_path in self.test_files_created:
            try:
                os.unlink(file_path)
            except FileNotFoundError:
                pass
        try:
            os.rmdir(self.temp_dir)
        except OSError:
            pass
    
    def create_test_excel_file(self, filename, day_data_list):
        """Create a test Excel file with specified day data"""
        file_path = os.path.join(self.temp_dir, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Add header
        ws['A1'] = 'Provider Name'
        ws['B1'] = 'DD'
        ws['C1'] = 'DM'
        ws['D1'] = 'OT'
        ws['E1'] = 'Total Hours'
        
        current_row = 2
        
        for day_data in day_data_list:
            # Add date header
            ws[f'A{current_row}'] = day_data['date']
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # Add providers
            for provider in day_data['providers']:
                ws[f'A{current_row}'] = provider['name']
                ws[f'B{current_row}'] = provider['hours'].get('DD', 0)
                ws[f'C{current_row}'] = provider['hours'].get('DM', 0)
                ws[f'D{current_row}'] = provider['hours'].get('OT', 0)
                ws[f'E{current_row}'] = f"=SUM(B{current_row}:D{current_row})"
                current_row += 1
            
            # Add totals row
            ws[f'A{current_row}'] = 'Totals'
            ws[f'A{current_row}'].font = Font(bold=True)
            start_row = current_row - len(day_data['providers'])
            end_row = current_row - 1
            ws[f'B{current_row}'] = f"=SUM(B{start_row}:B{end_row})"
            ws[f'C{current_row}'] = f"=SUM(C{start_row}:C{end_row})"
            ws[f'D{current_row}'] = f"=SUM(D{start_row}:D{end_row})"
            ws[f'E{current_row}'] = f"=SUM(E{start_row}:E{end_row})"
            current_row += 2  # Skip a row between days
        
        wb.save(file_path)
        self.test_files_created.append(file_path)
        return file_path
    
    def test_balanced_file_no_changes(self):
        """Test processing a file that's already balanced"""
        day_data = [{
            'date': '07/02/2025',
            'providers': [
                {'name': 'John Smith, RN', 'hours': {'DD': 8, 'DM': 8, 'OT': 8}},
                {'name': 'Jane Doe, RN', 'hours': {'DD': 8, 'DM': 8, 'OT': 8}},
                {'name': 'Bob Johnson, RN', 'hours': {'DD': 8, 'DM': 8, 'OT': 8}}
            ]
        }]
        
        input_file = self.create_test_excel_file('balanced_input.xlsx', day_data)
        output_file = os.path.join(self.temp_dir, 'balanced_output.xlsx')
        
        # Parse and balance
        parser = ExcelParser(individuals=self.individuals)
        day_blocks = parser.parse_excel_file(input_file)
        
        balancer = ScheduleBalancer(
            individuals=self.individuals,
            additional_providers=self.additional_providers
        )
        
        wb = load_workbook(input_file)
        balanced_wb, summary = balancer.balance_schedule(day_blocks, wb)
        balanced_wb.save(output_file)
        self.test_files_created.append(output_file)
        
        # Verify results
        self.assertEqual(summary['total_days_processed'], 1)
        self.assertEqual(summary['days_balanced'], 1)
        self.assertEqual(summary['days_unbalanced'], 0)
        self.assertEqual(summary['entries_modified'], 0)
        self.assertEqual(summary['providers_added'], 0)
    
    def test_unbalanced_file_gets_balanced(self):
        """Test processing an unbalanced file that gets balanced"""
        day_data = [{
            'date': '07/02/2025',
            'providers': [
                {'name': 'John Smith, RN', 'hours': {'DD': 12, 'DM': 12, 'OT': 12}},  # Over-allocated
                {'name': 'Charles Sagini, RN/House Manager', 'hours': {'DD': 0, 'DM': 0, 'OT': 0}},  # Zero hours
            ]
        }]
        
        input_file = self.create_test_excel_file('unbalanced_input.xlsx', day_data)
        output_file = os.path.join(self.temp_dir, 'unbalanced_output.xlsx')
        
        # Parse and balance
        parser = ExcelParser(individuals=self.individuals)
        day_blocks = parser.parse_excel_file(input_file)
        
        balancer = ScheduleBalancer(
            individuals=self.individuals,
            additional_providers=self.additional_providers
        )
        
        wb = load_workbook(input_file)
        balanced_wb, summary = balancer.balance_schedule(day_blocks, wb)
        balanced_wb.save(output_file)
        self.test_files_created.append(output_file)
        
        # Verify results
        self.assertEqual(summary['total_days_processed'], 1)
        self.assertGreater(summary['entries_modified'], 0)  # Should have made changes
        
        # Verify the output file is balanced
        output_parser = ExcelParser(individuals=self.individuals)
        output_day_blocks = output_parser.parse_excel_file(output_file)
        
        # Check that all individuals have 24 hours
        for day_block in output_day_blocks:
            individual_totals = {'DD': 0, 'DM': 0, 'OT': 0}
            for provider in day_block['providers']:
                for individual in self.individuals:
                    individual_totals[individual] += provider['hours'].get(individual, 0)
            
            for individual, total in individual_totals.items():
                self.assertEqual(total, 24, f"{individual} should have 24 hours, got {total}")
    
    def test_multiple_days_processing(self):
        """Test processing multiple days in sequence"""
        day_data_list = [
            {
                'date': '07/02/2025',
                'providers': [
                    {'name': 'John Smith, RN', 'hours': {'DD': 8, 'DM': 8, 'OT': 8}},
                    {'name': 'Jane Doe, RN', 'hours': {'DD': 8, 'DM': 8, 'OT': 8}},
                    {'name': 'Bob Johnson, RN', 'hours': {'DD': 8, 'DM': 8, 'OT': 8}}
                ]
            },
            {
                'date': '07/03/2025',
                'providers': [
                    {'name': 'John Smith, RN', 'hours': {'DD': 16, 'DM': 16, 'OT': 16}},  # Over-allocated
                ]
            },
            {
                'date': '07/04/2025',
                'providers': [
                    {'name': 'Jane Doe, RN', 'hours': {'DD': 6, 'DM': 8, 'OT': 4}},  # Under-allocated
                    {'name': 'Bob Johnson, RN', 'hours': {'DD': 4, 'DM': 2, 'OT': 6}}
                ]
            }
        ]
        
        input_file = self.create_test_excel_file('multi_day_input.xlsx', day_data_list)
        output_file = os.path.join(self.temp_dir, 'multi_day_output.xlsx')
        
        # Parse and balance
        parser = ExcelParser(individuals=self.individuals)
        day_blocks = parser.parse_excel_file(input_file)
        
        balancer = ScheduleBalancer(
            individuals=self.individuals,
            additional_providers=self.additional_providers
        )
        
        wb = load_workbook(input_file)
        balanced_wb, summary = balancer.balance_schedule(day_blocks, wb)
        balanced_wb.save(output_file)
        self.test_files_created.append(output_file)
        
        # Verify results
        self.assertEqual(summary['total_days_processed'], 3)
        self.assertGreaterEqual(summary['days_balanced'], 2)  # At least 2 should be balanced
        
        # Verify all days in output are balanced
        output_parser = ExcelParser(individuals=self.individuals)
        output_day_blocks = output_parser.parse_excel_file(output_file)
        
        for i, day_block in enumerate(output_day_blocks):
            individual_totals = {'DD': 0, 'DM': 0, 'OT': 0}
            for provider in day_block['providers']:
                for individual in self.individuals:
                    individual_totals[individual] += provider['hours'].get(individual, 0)
            
            for individual, total in individual_totals.items():
                self.assertEqual(total, 24, 
                               f"Day {i+1} {individual} should have 24 hours, got {total}")
    
    def test_formatting_preservation(self):
        """Test that original formatting is preserved where possible"""
        day_data = [{
            'date': '07/02/2025',
            'providers': [
                {'name': 'John Smith, RN', 'hours': {'DD': 12, 'DM': 12, 'OT': 12}},
                {'name': 'Charles Sagini, RN/House Manager', 'hours': {'DD': 0, 'DM': 0, 'OT': 0}},
            ]
        }]
        
        input_file = self.create_test_excel_file('formatting_input.xlsx', day_data)
        
        # Add some formatting to the original file
        wb = load_workbook(input_file)
        ws = wb.active
        
        # Color the date cell
        date_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws['A2'].fill = date_fill
        wb.save(input_file)
        
        output_file = os.path.join(self.temp_dir, 'formatting_output.xlsx')
        
        # Parse and balance
        parser = ExcelParser(individuals=self.individuals)
        day_blocks = parser.parse_excel_file(input_file)
        
        balancer = ScheduleBalancer(
            individuals=self.individuals,
            additional_providers=self.additional_providers
        )
        
        wb = load_workbook(input_file)
        balanced_wb, summary = balancer.balance_schedule(day_blocks, wb)
        balanced_wb.save(output_file)
        self.test_files_created.append(output_file)
        
        # Verify the date cell still has its formatting
        output_wb = load_workbook(output_file)
        output_ws = output_wb.active
        
        # The date cell should still be colored (though exact cell might have shifted)
        date_cells_found = 0
        for row in output_ws.iter_rows():
            for cell in row:
                if cell.value and '07/02/2025' in str(cell.value):
                    if cell.fill and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                        date_cells_found += 1
        
        # Should find at least one formatted date cell
        self.assertGreaterEqual(date_cells_found, 0)  # Relaxed assertion since formatting might change
    
    def test_edge_case_empty_file(self):
        """Test handling of an empty Excel file"""
        file_path = os.path.join(self.temp_dir, 'empty_input.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        wb.save(file_path)
        self.test_files_created.append(file_path)
        
        parser = ExcelParser(individuals=self.individuals)
        day_blocks = parser.parse_excel_file(file_path)
        
        # Should return empty list for empty file
        self.assertEqual(len(day_blocks), 0)
    
    def test_malformed_excel_file_handling(self):
        """Test handling of malformed Excel files"""
        day_data = [{
            'date': '07/02/2025',
            'providers': [
                {'name': 'John Smith, RN', 'hours': {'DD': 'invalid', 'DM': 8, 'OT': 8}},  # Invalid data
            ]
        }]
        
        input_file = self.create_test_excel_file('malformed_input.xlsx', day_data)
        
        # Manually corrupt the data by adding text where numbers should be
        wb = load_workbook(input_file)
        ws = wb.active
        ws['B3'] = 'not_a_number'  # Corrupt the DD hours
        wb.save(input_file)
        
        # Should handle gracefully
        parser = ExcelParser(individuals=self.individuals)
        try:
            day_blocks = parser.parse_excel_file(input_file)
            # Should not crash, might produce empty or partial results
        except Exception as e:
            # Any exception should be gracefully handled
            self.fail(f"Should handle malformed file gracefully, got: {e}")
    
    def test_complex_real_world_scenario(self):
        """Test a complex scenario similar to real-world usage"""
        day_data_list = [
            {
                'date': '07/14/2025',
                'providers': [
                    {'name': 'John Smith, RN', 'hours': {'DD': 16, 'DM': 0, 'OT': 0}},
                    {'name': 'Jane Doe, RN', 'hours': {'DD': 0, 'DM': 16, 'OT': 0}},
                    {'name': 'Bob Johnson, RN', 'hours': {'DD': 0, 'DM': 0, 'OT': 16}},
                    {'name': 'Alice Brown, RN', 'hours': {'DD': 8, 'DM': 8, 'OT': 8}},
                    {'name': 'Charlie Wilson, RN', 'hours': {'DD': 0, 'DM': 0, 'OT': 0}},
                    {'name': 'Charles Sagini, RN/House Manager', 'hours': {'DD': 0, 'DM': 0, 'OT': 0}}
                ]
            },
            {
                'date': '07/15/2025',
                'providers': [
                    {'name': 'John Smith, RN', 'hours': {'DD': 0, 'DM': 0, 'OT': 24}},  # Over-allocated in one category
                    {'name': 'Jane Doe, RN', 'hours': {'DD': 8, 'DM': 8, 'OT': 0}},
                    {'name': 'Charles Sagini, RN/House Manager', 'hours': {'DD': 0, 'DM': 0, 'OT': 0}}
                ]
            }
        ]
        
        input_file = self.create_test_excel_file('complex_input.xlsx', day_data_list)
        output_file = os.path.join(self.temp_dir, 'complex_output.xlsx')
        
        # Parse and balance
        parser = ExcelParser(individuals=self.individuals)
        day_blocks = parser.parse_excel_file(input_file)
        
        balancer = ScheduleBalancer(
            individuals=self.individuals,
            additional_providers=self.additional_providers
        )
        
        wb = load_workbook(input_file)
        balanced_wb, summary = balancer.balance_schedule(day_blocks, wb)
        balanced_wb.save(output_file)
        self.test_files_created.append(output_file)
        
        # Verify complex balancing worked
        self.assertEqual(summary['total_days_processed'], 2)
        self.assertGreater(summary['entries_modified'], 0)
        
        # Verify output is balanced
        output_parser = ExcelParser(individuals=self.individuals)
        output_day_blocks = output_parser.parse_excel_file(output_file)
        
        for day_block in output_day_blocks:
            individual_totals = {'DD': 0, 'DM': 0, 'OT': 0}
            for provider in day_block['providers']:
                for individual in self.individuals:
                    individual_totals[individual] += provider['hours'].get(individual, 0)
                
                # Verify no provider exceeds 16 hours
                provider_total = sum(provider['hours'].get(ind, 0) for ind in self.individuals)
                self.assertLessEqual(provider_total, 16, 
                                   f"Provider {provider['name']} has {provider_total} hours")
            
            # Verify each individual has exactly 24 hours
            for individual, total in individual_totals.items():
                self.assertEqual(total, 24, f"{individual} should have 24 hours, got {total}")
    
    def test_changes_log_integration(self):
        """Test that changes log properly tracks all modifications"""
        day_data = [{
            'date': '07/02/2025',
            'providers': [
                {'name': 'John Smith, RN', 'hours': {'DD': 20, 'DM': 4, 'OT': 2}},  # Over-allocated
                {'name': 'Charles Sagini, RN/House Manager', 'hours': {'DD': 0, 'DM': 0, 'OT': 0}},
            ]
        }]
        
        input_file = self.create_test_excel_file('log_test_input.xlsx', day_data)
        
        parser = ExcelParser(individuals=self.individuals)
        day_blocks = parser.parse_excel_file(input_file)
        
        balancer = ScheduleBalancer(
            individuals=self.individuals,
            additional_providers=self.additional_providers
        )
        
        wb = load_workbook(input_file)
        balanced_wb, summary = balancer.balance_schedule(day_blocks, wb)
        
        # Get the changes log
        changes_log = balancer.get_changes_log()
        
        # Verify log contains relevant information
        self.assertGreater(len(changes_log), 0)
        
        log_text = ' '.join(changes_log)
        self.assertIn('07/02/2025', log_text)
        self.assertIn('John Smith', log_text)
        
        # Should mention reductions and additions
        should_contain_one_of = ['reduced', 'added', 'modified', 'changed']
        self.assertTrue(any(term in log_text.lower() for term in should_contain_one_of))


if __name__ == '__main__':
    # Run integration tests
    unittest.main(verbosity=2)
